from datetime import datetime, timedelta
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from fpdf import FPDF
from tkinter import messagebox





def save_to_excel(filename, data, headers):
    # Δημιουργία ή φόρτωση αρχείου Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Data'

    # Προσθήκη κεφαλίδων
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Προσθήκη δεδομένων
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=value)

    # Αποθήκευση αρχείου
    workbook.save(filename)






def find_signatures_and_worktime_diffs(records, employees, min_hours=7, min_minutes=45):
    no_signatures = []
    one_signature = []
    less_than_min_worktime = []
    present_employees = set()

    for employee_id, record in records.items():
        entry_time = record.get("entry_time")
        exit_time = record.get("exit_time")
        entry_signature_url = record.get("entry_signature_url")
        exit_signature_url = record.get("exit_signature_url")

        # Ελέγχουμε ποιοι υπάλληλοι έχουν καταχωρήσεις
        present_employees.add(employee_id)

        # Έλεγχος για υπογραφές
        signatures_count = sum(bool(url) for url in [entry_signature_url, exit_signature_url])



        if signatures_count == 1:
            one_signature.append((employee_id, record["name"]))
            print(f"Μία υπογραφή: {employee_id}, {record['name']}")

        # Έλεγχος για διαφορά ώρας εξόδου - εισόδου
        if entry_time and exit_time:
            try:
                # Καθαρισμός της ώρας
                entry_time_clean = entry_time.strip()
                exit_time_clean = exit_time.strip()

                # Χρήση του σωστού format με δευτερόλεπτα
                entry_time_obj = datetime.strptime(entry_time_clean, '%H:%M:%S')
                exit_time_obj = datetime.strptime(exit_time_clean, '%H:%M:%S')
                work_duration = exit_time_obj - entry_time_obj

                # Αν η διαφορά είναι μικρότερη από το κατώφλι
                if work_duration < timedelta(hours=min_hours, minutes=min_minutes):
                    less_than_min_worktime.append((employee_id, record["name"], str(work_duration)))
                    print(f"Διαφορά μικρότερη από 7:45: {employee_id}, {record['name']}")

            except ValueError:
                print(f"Μη έγκυρη ώρα για τον υπάλληλο {employee_id}: {entry_time} - {exit_time}")
                continue

    # Εύρεση απόντων υπαλλήλων (αυτοί που δεν υπάρχουν καθόλου στα records)
    no_signatures = [(employee_id, employees[employee_id]) for employee_id in employees if
                        employee_id not in present_employees]
    print(f"Υπάλληλοι χωρίς καμία καταχώρηση (απόντες): {len(no_signatures)}")
    print(present_employees)
    # Επιστροφή των αποτελεσμάτων
    return no_signatures, one_signature, less_than_min_worktime


def generate_report_pdf_and_excel(date):
    # Έλεγχος αν υπάρχουν τα αρχεία υπαλλήλων και δεδομένων
    if not os.path.exists('employees_list.xlsx'):
        messagebox.showerror("Σφάλμα", "Το αρχείο 'employees_list.xlsx' δεν βρέθηκε.")
        return

    if not os.path.exists(f'employee_data_{date}.xlsx'):
        messagebox.showerror("Σφάλμα", f"Το αρχείο 'employee_data_{date}.xlsx' δεν βρέθηκε.")
        return

    # Φόρτωση δεδομένων από τα αρχεία Excel
    employees_workbook = load_workbook('employees_list.xlsx')
    employees_sheet = employees_workbook.active
    employee_data_workbook = load_workbook(f'employee_data_{date}.xlsx')
    employee_data_sheet = employee_data_workbook.active

    # Φόρτωση λίστας υπαλλήλων από το αρχείο "employees_list.xlsx"
    employees = {}
    for row in employees_sheet.iter_rows(min_row=2, values_only=True):
        employee_id, name = str(row[0]).strip(), str(row[1]).strip() if row[1] else ""
        name = name.capitalize()
        employees[employee_id] = name

    # Ταξινόμηση των υπαλλήλων με βάση το ονοματεπώνυμο
    employees = dict(sorted(employees.items(), key=lambda item: item[1]))

    # Φόρτωση δεδομένων καταχωρήσεων από το αρχείο "employee_data_{date}.xlsx"
    records_by_date = {}
    for row in employee_data_sheet.iter_rows(min_row=2, values_only=True):
        employee_id, name, date, entry_time, entry_signature_url, exit_time, exit_signature_url = row
        if date not in records_by_date:
            records_by_date[date] = {}

        records_by_date[date][employee_id] = {
            "name": name,
            "entry_time": entry_time,
            "entry_signature_url": entry_signature_url,
            "exit_time": exit_time,
            "exit_signature_url": exit_signature_url
        }

    # Δημιουργία του PDF
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Προσθήκη και χρήση της DejaVuSans γραμματοσειράς
    pdf.add_font('DejaVu', '', 'DejaVuSans.ttf', uni=True)
    pdf.add_font('DejaVu-Bold', '', 'DejaVuSans-Bold.ttf', uni=True)
    pdf.set_font('DejaVu-Bold', '', 14)
    pdf.cell(200, 10, text="Αναφορά Υπαλλήλων", new_x='LMARGIN', new_y='NEXT', align='C')

    # Επεξεργασία κάθε ημερομηνίας
    for date, records in sorted(records_by_date.items()):
        pdf.set_font('DejaVu-Bold', '', 12)
        pdf.cell(200, 10, text=f"Ημερομηνία: {date.strftime('%d/%m/%Y')}", new_x='LMARGIN', new_y='NEXT')

        # Σχεδίαση της επικεφαλίδας του πίνακα
        pdf.set_font('DejaVu-Bold', '', 10)
        #pdf.cell(40, 10, "Αριθμός Μητρώου", border=1)
        pdf.cell(60, 10, "Ονοματεπώνυμο", border=1)
        pdf.cell(30, 10, "Ώρα Εισόδου", border=1)
        pdf.cell(30, 10, "Ώρα Εξόδου", border=1)
        pdf.cell(20, 10, "Είσοδος", border=1)
        pdf.cell(20, 10, "Έξοδος", border=1)
        pdf.ln()

        # Δημιουργία λιστών για αποθήκευση σε Excel
        no_signatures, one_signature, less_than_worktime  = find_signatures_and_worktime_diffs(records,employees)
        # Προσθήκη δεδομένων υπαλλήλων στο PDF
        for employee_id, employee_name in employees.items():
            pdf.set_font('DejaVu', '', 10)
            #pdf.cell(40, 10, employee_id, border=1)
            pdf.cell(60, 10, employee_name, border=1)

            if employee_id in records:
                record = records[employee_id]
                entry_time = record.get("entry_time", "")
                exit_time = record.get("exit_time", "")
                entry_signature_url = record.get("entry_signature_url")
                exit_signature_url = record.get("exit_signature_url")

                pdf.cell(30, 10, str(entry_time) if entry_time else "-", border=1)
                pdf.cell(30, 10, str(exit_time) if exit_time else "-", border=1)

                # Υπογραφές
                if entry_signature_url and os.path.exists(entry_signature_url):
                    pdf.cell(20, 10, "", border=1)
                    pdf.image(entry_signature_url, x=pdf.get_x() - 20, y=pdf.get_y(), w=15, h=10)
                else:
                    pdf.cell(20, 10, "Απών", border=1)

                if exit_signature_url and os.path.exists(exit_signature_url):
                    pdf.cell(20, 10, "", border=1)
                    pdf.image(exit_signature_url, x=pdf.get_x() - 20, y=pdf.get_y(), w=15, h=10)
                else:
                    pdf.cell(20, 10, "Απών", border=1)

            else:
                pdf.cell(30, 10, "-", border=1)
                pdf.cell(30, 10, "-", border=1)
                pdf.cell(20, 10, "Απών", border=1)
                pdf.cell(20, 10, "Απών", border=1)
            pdf.ln()

        # Αποθήκευση δεδομένων σε Excel
        save_to_excel(f"no_signatures_{date.strftime('%Y-%m-%d')}.xlsx", no_signatures,
                      ["Αριθμός Μητρώου", "Ονοματεπώνυμο"])
        save_to_excel(f"one_signature_{date.strftime('%Y-%m-%d')}.xlsx", one_signature,
                      ["Αριθμός Μητρώου", "Ονοματεπώνυμο"])
        save_to_excel(f"less_than_worktime_{date.strftime('%Y-%m-%d')}.xlsx", less_than_worktime,
                      ["Αριθμός Μητρώου", "Ονοματεπώνυμο", "Διάρκεια Εργασίας"])

    pdf_filename = f"employees_report_{date.strftime('%Y-%m-%d')}.pdf"
    pdf.output(pdf_filename)

    messagebox.showinfo("Επιτυχία", f"Το PDF '{pdf_filename}' και τα αρχεία Excel δημιουργήθηκαν με επιτυχία.")



def generate_report_pdf_and_excel_with_date(delay,merged_file):
    # Έλεγχος αν υπάρχουν τα αρχεία υπαλλήλων και δεδομένων
    if not os.path.exists('employees_list.xlsx'):
        messagebox.showerror("Σφάλμα", "Το αρχείο 'employees_list.xlsx' δεν βρέθηκε.")
        return

    if not os.path.exists(merged_file):
        messagebox.showerror("Σφάλμα", f"Το αρχείο '{merged_file}' δεν βρέθηκε.")
        return

    # Φόρτωση δεδομένων από τα αρχεία Excel
    employees_workbook = load_workbook('employees_list.xlsx')
    employees_sheet = employees_workbook.active
    employee_data_workbook = load_workbook(f'{merged_file}')
    employee_data_sheet = employee_data_workbook.active

    employees = {}
    for row in employees_sheet.iter_rows(min_row=2, values_only=True):
        employee_id, name = str(row[0]).strip(), str(row[1]).strip() if row[1] else ""
        name = name.capitalize()
        employees[employee_id] = name

    # Ταξινόμηση των υπαλλήλων με βάση το ονοματεπώνυμο
    employees = dict(sorted(employees.items(), key=lambda item: item[1]))

    # Φόρτωση δεδομένων καταχωρήσεων από το αρχείο "employee_data_{date}.xlsx"
    records_by_date = {}
    for row in employee_data_sheet.iter_rows(min_row=2, values_only=True):
        employee_id, name, date, entry_time, entry_signature_url, exit_time, exit_signature_url = row
        if date not in records_by_date:
            records_by_date[date] = {}

        records_by_date[date][employee_id] = {
            "name": name,
            "entry_time": entry_time,
            "entry_signature_url": entry_signature_url,
            "exit_time": exit_time,
            "exit_signature_url": exit_signature_url
        }



    # Επεξεργασία κάθε ημερομηνίας
    for date, records in sorted(records_by_date.items()):
        # Δημιουργία του PDF
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # Προσθήκη και χρήση της DejaVuSans γραμματοσειράς
        pdf.add_font('DejaVu', '', 'DejaVuSans.ttf', uni=True)
        pdf.add_font('DejaVu-Bold', '', 'DejaVuSans-Bold.ttf', uni=True)
        pdf.set_font('DejaVu-Bold', '', 14)
        pdf.cell(200, 10, text="Αναφορά Υπαλλήλων", new_x='LMARGIN', new_y='NEXT', align='C')
        pdf.set_font('DejaVu-Bold', '', 12)
        pdf.cell(200, 10, text=f"Ημερομηνία: {date.strftime('%d/%m/%Y')}", new_x='LMARGIN', new_y='NEXT')

        # Σχεδίαση της επικεφαλίδας του πίνακα
        pdf.set_font('DejaVu-Bold', '', 10)
        pdf.cell(60, 10, "Ονοματεπώνυμο", border=1)
        pdf.cell(30, 10, "Ώρα Εισόδου", border=1)
        pdf.cell(30, 10, "Ώρα Εξόδου", border=1)
        pdf.cell(20, 10, "Είσοδος", border=1)
        pdf.cell(20, 10, "Έξοδος", border=1)
        pdf.ln()

        # Δημιουργία λιστών για αποθήκευση σε Excel
        no_signatures, one_signature, less_than_worktime = find_signatures_and_worktime_diffs(records, employees,min_minutes=60-delay)

        # Προσθήκη δεδομένων υπαλλήλων στο PDF
        for employee_id, employee_name in employees.items():
            pdf.set_font('DejaVu', '', 10)
            pdf.cell(60, 10, employee_name, border=1)

            if employee_id in records:
                record = records[employee_id]
                entry_time = record.get("entry_time", "")
                exit_time = record.get("exit_time", "")
                entry_signature_url = record.get("entry_signature_url")
                exit_signature_url = record.get("exit_signature_url")

                pdf.cell(30, 10, str(entry_time) if entry_time else "-", border=1)
                pdf.cell(30, 10, str(exit_time) if exit_time else "-", border=1)

                # Υπογραφές
                if entry_signature_url and os.path.exists(entry_signature_url):
                    pdf.cell(20, 10, "", border=1)
                    pdf.image(entry_signature_url, x=pdf.get_x() - 20, y=pdf.get_y(), w=15, h=10)
                else:
                    pdf.cell(20, 10, "Απών", border=1)

                if exit_signature_url and os.path.exists(exit_signature_url):
                    pdf.cell(20, 10, "", border=1)
                    pdf.image(exit_signature_url, x=pdf.get_x() - 20, y=pdf.get_y(), w=15, h=10)
                else:
                    pdf.cell(20, 10, "Απών", border=1)

            else:
                pdf.cell(30, 10, "-", border=1)
                pdf.cell(30, 10, "-", border=1)
                pdf.cell(20, 10, "Απών", border=1)
                pdf.cell(20, 10, "Απών", border=1)
            pdf.ln()

        # Λήψη του έτους και του μήνα από την ημερομηνία
        year_str = date.strftime('%Y')
        month_str = date.strftime('%m')
        date_str = date.strftime('%d-%m-%Y')

        # Δημιουργία φακέλων Έτος/Μήνας
        folder_name = f"Reports/{year_str}/{month_str}/Report_{date_str}"
        if not os.path.exists(folder_name):
            os.makedirs(folder_name)
        # Αποθήκευση Excel αρχείου με δεδομένα της ημέρας στον φάκελο
        if date in records_by_date:  # Έλεγχος αν υπάρχουν εγγραφές για τη συγκεκριμένη ημερομηνία
            records_for_date = []
            for employee_id, record in records_by_date[date].items():
                records_for_date.append([
                    employee_id,
                    record.get("name", ""),
                    date.strftime('%d/%m/%Y'),
                    record.get("entry_time", ""),
                    record.get("entry_signature_url", ""),
                    record.get("exit_time", ""),
                    record.get("exit_signature_url", "")
                ])

            # Κλήση της save_to_excel με τα δεδομένα της ημέρας
            save_to_excel(f"{folder_name}/employee_data_{date_str}.xlsx", records_for_date,
                          ["Αριθμός Μητρώου", "Ονοματεπώνυμο", "Ημερομηνία", "Ώρα εισόδου", "URL Υπογραφής εισόδου",
                           "Ώρα εξόδου", "URL Υπογραφής εξόδου"])

        # Αποθήκευση PDF και Excel ανά ημερομηνία στον φάκελο
        save_to_excel(f"{folder_name}/no_signatures_{date_str}.xlsx", no_signatures,
                      ["Αριθμός Μητρώου", "Ονοματεπώνυμο"])
        save_to_excel(f"{folder_name}/one_signature_{date_str}.xlsx", one_signature,
                      ["Αριθμός Μητρώου", "Ονοματεπώνυμο"])
        save_to_excel(f"{folder_name}/less_than_worktime_{date_str}.xlsx", less_than_worktime,
                      ["Αριθμός Μητρώου", "Ονοματεπώνυμο", "Διάρκεια Εργασίας"])
        # Προσθήκη αθροιστικού πίνακα στο τέλος του PDF
        pdf.add_page()
        pdf.set_font('DejaVu-Bold', '', 12)
        pdf.cell(200, 10, text="Αθροιστικά Στοιχεία", new_x='LMARGIN', new_y='NEXT')

        total_employees = len(employees)
        pdf.cell(200, 10, text=f"Σύνολο Εργαζομένων: {total_employees}", new_x='LMARGIN', new_y='NEXT')
        pdf.cell(200, 10, text=f"Απόντες: {len(no_signatures)}", new_x='LMARGIN', new_y='NEXT')
        pdf.cell(200, 10, text=f"Μία Υπογραφή: {len(one_signature)}", new_x='LMARGIN', new_y='NEXT')
        pdf.cell(200, 10, text=f"Καθυστερήσανε: {len(less_than_worktime)}", new_x='LMARGIN', new_y='NEXT')
        pdf_filename = f"{folder_name}/employees_report_{date_str}.pdf"
        pdf.output(pdf_filename)
        messagebox.showinfo("Επιτυχία", f"Το PDF employees_report_{date_str}.pdf και τα αρχεία Excel δημιουργήθηκαν με επιτυχία στους φακέλους τους.")
