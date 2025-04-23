import sys
import ctypes
from fpdf import FPDF
import re
# Enable DPI awareness for accurate screen scaling
ctypes.windll.user32.SetProcessDPIAware()
import tkinter as tk
from tkinter import messagebox, Frame
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
from PIL import ImageGrab
import make_Data

def generate_pdf(date):
    # Έλεγχος αν υπάρχουν τα αρχεία υπαλλήλων και δεδομένων
    if not os.path.exists('employees_list.xlsx'):
        messagebox.showerror("Σφάλμα", "Το αρχείο 'employees_list.xlsx' δεν βρέθηκε.")
        return

    if not os.path.exists(f'employee_data_{date}.xlsx'):
        messagebox.showerror("Σφάλμα", f"Το αρχείο 'employee_data_{date}.xlsx' δεν βρέθηκε.")
        return

    # Φόρτωση δεδομένων από τα αρχεία Excel
    employees_workbook = load_workbook('employees_list.xlsx')
    employees_sheet = employees_workbook.active
    employee_data_workbook = load_workbook(f'employee_data_{date}.xlsx')
    employee_data_sheet = employee_data_workbook.active

    # Φόρτωση λίστας υπαλλήλων από το αρχείο "employees_list.xlsx"
    employees = {}
    for row in employees_sheet.iter_rows(min_row=2, values_only=True):
        employee_id, name = str(row[0]).strip(), str(row[1]).strip() if row[1] else ""
        # Μετατροπή του πρώτου γράμματος σε κεφαλαίο
        name = name.capitalize()
        employees[employee_id] = name

    # Ταξινόμηση των υπαλλήλων με βάση το ονοματεπώνυμο
    employees = dict(sorted(employees.items(), key=lambda item: item[1]))

    # Φόρτωση δεδομένων καταχωρήσεων από το αρχείο "employee_data_{date}.xlsx"
    records_by_date = {}
    for row in employee_data_sheet.iter_rows(min_row=2, values_only=True):
        employee_id, name, date, entry_time, entry_signature_url, exit_time, exit_signature_url = row
        if date not in records_by_date:
            records_by_date[date] = {}

        # Αποθήκευση της καταχώρησης στη συγκεκριμένη ημερομηνία για τον συγκεκριμένο υπάλληλο
        records_by_date[date][employee_id] = {
            "name": name,
            "entry_time": entry_time,
            "entry_signature_url": entry_signature_url,
            "exit_time": exit_time,
            "exit_signature_url": exit_signature_url
        }

    # Δημιουργία του PDF με χρήση της fpdf2
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Προσθήκη και χρήση της DejaVuSans γραμματοσειράς που υποστηρίζει ελληνικά
    pdf.add_font('DejaVu', '', 'DejaVuSans.ttf', uni=True)
    pdf.add_font('DejaVu-Bold', '', 'DejaVuSans-Bold.ttf', uni=True)

    # Χρήση της γραμματοσειράς για ελληνικούς χαρακτήρες
    pdf.set_font('DejaVu-Bold', '', 14)  # Χρήση για τίτλους
    pdf.cell(200, 10, text="Αναφορά Υπαλλήλων", new_x='LMARGIN', new_y='NEXT', align='C')

    # Διέλευση όλων των ημερομηνιών
    for date, records in sorted(records_by_date.items()):
        pdf.set_font('DejaVu-Bold', '', 12)
        pdf.cell(200, 10, text=f"Ημερομηνία: {date.strftime('%d/%m/%Y')}", new_x='LMARGIN', new_y='NEXT')

        # Σχεδίαση της επικεφαλίδας του πίνακα
        pdf.set_font('DejaVu-Bold', '', 10)
        pdf.cell(40, 10, "Αριθμός Μητρώου", border=1)
        pdf.cell(50, 10, "Ονοματεπώνυμο", border=1)
        pdf.cell(30, 10, "Ώρα Εισόδου", border=1)
        pdf.cell(30, 10, "Ώρα Εξόδου", border=1)
        pdf.cell(20, 10, "Υπογραφή Εισόδου", border=1)
        pdf.cell(20, 10, "Υπογραφή Εξόδου", border=1)
        pdf.ln()

        # Δημιουργία λίστας υπαλλήλων που δεν υπέγραψαν
        missing_signatures = []

        # Διέλευση όλων των υπαλλήλων από τη λίστα
        for employee_id, employee_name in employees.items():
            pdf.set_font('DejaVu', '', 10)
            pdf.cell(40, 10, employee_id, border=1)
            pdf.cell(50, 10, employee_name, border=1)

            # Προσθήκη των ωρών εισόδου και εξόδου
            if employee_id in records:
                record = records[employee_id]
                entry_time = record.get("entry_time", "")
                exit_time = record.get("exit_time", "")
                entry_signature_url = record.get("entry_signature_url")
                exit_signature_url = record.get("exit_signature_url")

                pdf.cell(30, 10, str(entry_time) if entry_time else "-", border=1)
                pdf.cell(30, 10, str(exit_time) if exit_time else "-", border=1)

                # Προσθήκη της εικόνας της υπογραφής εισόδου αν υπάρχει
                if entry_signature_url and os.path.exists(entry_signature_url):
                    pdf.cell(20, 10, "", border=1)
                    pdf.image(entry_signature_url, x=pdf.get_x() - 20, y=pdf.get_y(), w=15, h=10)
                else:
                    pdf.cell(20, 10, "Απών", border=1)

                # Προσθήκη της εικόνας της υπογραφής εξόδου αν υπάρχει
                if exit_signature_url and os.path.exists(exit_signature_url):
                    pdf.cell(20, 10, "", border=1)
                    pdf.image(exit_signature_url, x=pdf.get_x() - 20, y=pdf.get_y(), w=15, h=10)
                else:
                    pdf.cell(20, 10, "Απών", border=1)

            else:
                pdf.cell(30, 10, "-", border=1)
                pdf.cell(30, 10, "-", border=1)
                pdf.cell(20, 10, "Απών", border=1)
                pdf.cell(20, 10, "Απών", border=1)
                # Προσθήκη του υπαλλήλου στη λίστα "χωρίς υπογραφή"
                missing_signatures.append((employee_id, employee_name))

            pdf.ln()

        # Προσθήκη πίνακα υπαλλήλων που δεν υπέγραψαν για τη συγκεκριμένη ημερομηνία
        if missing_signatures:
            pdf.set_font('DejaVu-Bold', '', 12)
            pdf.cell(200, 10, txt=f"Υπάλληλοι Χωρίς Υπογραφή για την Ημερομηνία: {date.strftime('%d/%m/%Y')}", new_x='LMARGIN', new_y='NEXT')

            # Σχεδίαση της επικεφαλίδας του πίνακα για "χωρίς υπογραφή"
            pdf.set_font('DejaVu-Bold', '', 10)
            pdf.cell(40, 10, "Αριθμός Μητρώου", border=1)
            pdf.cell(100, 10, "Ονοματεπώνυμο", border=1)
            pdf.ln()

            pdf.set_font('DejaVu', '', 10)
            for employee_id, employee_name in missing_signatures:
                pdf.cell(40, 10, employee_id, border=1)
                pdf.cell(100, 10, employee_name, border=1)
                pdf.ln()

        pdf.ln()
    # Στο σημείο όπου δημιουργείται το όνομα του αρχείου
    pdf_filename = f"employees_report_{date.strftime('%Y-%m-%d')}.pdf"
    pdf.output(pdf_filename)


    messagebox.showinfo("Επιτυχία", f"Το PDF '{pdf_filename}' δημιουργήθηκε με επιτυχία.")


def save_signature_image(employee_id,canvas,state):
    # Ορισμός του ονόματος του αρχείου με βάση το ΑΜ και την ημερομηνία
    current_date_str = datetime.now().strftime("%Y%m%d")
    filename = f"signature_{employee_id}_{current_date_str}_{state}.png"
    '''
    # Ορισμός του μονοπατιού του φακέλου "images"
    images_folder = os.path.join(os.getcwd(), "images")

    # Δημιουργία του φακέλου "images" αν δεν υπάρχει
    if not os.path.exists(images_folder):
        os.makedirs(images_folder)

    # Ορισμός του πλήρους μονοπατιού για την αποθήκευση του αρχείου
    filepath = os.path.join(images_folder, filename)
    '''
    # Ορισμός του μονοπατιού του φακέλου "images" ως σχετική διαδρομή
    images_folder = "images"

    # Δημιουργία του φακέλου "images" αν δεν υπάρχει
    if not os.path.exists(images_folder):
        os.makedirs(images_folder)


    filepath = os.path.join(images_folder, filename)  # Αποθηκεύεται η σχετική διαδρομή
    # Έλεγχος αν ο καμβάς είναι άδειος (δηλ. αν δεν υπάρχουν στοιχεία στον καμβά)
    if len(canvas.find_all()) == 0:
        messagebox.showwarning("Υπογραφή Απαιτείται",
                               "Παρακαλώ προσθέστε την υπογραφή σας στον καμβά πριν αποθηκεύσετε.")
        return None  # Επιστροφή None για να υποδείξουμε ότι η αποθήκευση απέτυχε

    # Λήψη των συντεταγμένων του καμβά και αποθήκευση της εικόνας
    canvas.update_idletasks()  # Ensure the canvas is updated with the latest dimensions

    # Υπολογισμός συντεταγμένων του καμβά στην οθόνη
    x = canvas.winfo_rootx()
    y = canvas.winfo_rooty()
    width = canvas.winfo_width()
    height = canvas.winfo_height()

    # Εκτύπωση για debugging
    print(f"Canvas coordinates (screen): ({x}, {y})")
    print(f"Canvas size: {width}x{height}")

    # Υπολογισμός του κάτω δεξιού σημείου του καμβά
    x1 = x + width
    y1 = y + height

    # Χρήση bbox για λήψη συγκεκριμένης περιοχής της οθόνης
    bbox = (x, y, x1, y1)
    captured_image = ImageGrab.grab(bbox=bbox)

    # Αποθήκευση της εικόνας στο φάκελο "images"
    captured_image.save(filepath)

    # Επιστροφή του μονοπατιού αρχείου για αποθήκευση στο Excel
    return filepath


# Καθαρισμός του καμβά
def clear_canvas(canvas):
    canvas.delete("all")

def save_employee_data(canvas, state):
    employee_id = None
    if state == "Είσοδος":
        employee_id = str(entry.get().strip())  # Αφαιρούμε τυχόν κενά διαστήματα και το μετατρέπουμε σε string

    if state == "Έξοδος":
        employee_id = str(exit_entry.get().strip())  # Αφαιρούμε τυχόν κενά διαστήματα και το μετατρέπουμε σε string

    if employee_id == "":
        messagebox.showwarning("Εισαγωγή Δεδομένων", "Παρακαλώ εισάγετε έναν αριθμό μητρώου.")
        clear_canvas(canvas)
        return

    # Έλεγχος αν το employee_id ταιριάζει με το πρότυπο "pdf" ακολουθούμενο από ημερομηνία
    match = re.match(r'^pdf(\d{2}-\d{2}-\d{4})$', employee_id)
    if match:
        # Αποθήκευση της ημερομηνίας στη μεταβλητή data
        data = match.group(1)
        # Δημιούργησε το pdf
        #generate_pdf(data)
        make_Data.generate_report_pdf_and_excel(data)
        clear_canvas(canvas)
        return

    # Έλεγχος αν ο αριθμός μητρώου υπάρχει στη λίστα υπαλλήλων employees
    employee = next((emp for emp in employees if str(emp['id']).strip() == employee_id), None)
    entry_time = None
    exit_time = None

    if employee:
        # Παίρνουμε την τρέχουσα ημερομηνία και ώρα
        current_date = datetime.now().date()
        current_time = datetime.now().strftime("%H:%M:%S")
        # Μετατροπή της ημερομηνίας σε string με τη μορφή YYYY-MM-DD
        date_str = current_date.strftime("%d-%m-%Y")
        # Έλεγχος αν υπάρχει το αρχείο καταχώρησης
        if os.path.exists(f'employee_data_{date_str}.xlsx'):
            workbook = load_workbook(f'employee_data_{date_str}.xlsx')
            sheet = workbook.active
        else:
            # Δημιουργία νέου Excel αρχείου αν δεν υπάρχει
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Αριθμός Μητρώου", "Ονοματεπώνυμο", "Ημερομηνία", "Ώρα εισόδου", "URL Υπογραφής εισόδου", "Ώρα εξόδου", "URL Υπογραφής εξόδου"])



        # Έλεγχος εάν ο υπάλληλος έχει ήδη καταχωρηθεί σήμερα
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            employee_id_from_file = str(row[0])
            date_from_file = row[2].date() if isinstance(row[2], datetime) else row[2]
            entry_time = row[3]
            exit_time = row[5]

            # Έλεγχος εάν το ID και η ημερομηνία ταυτίζονται με την τρέχουσα ημερομηνία
            if employee_id_from_file == employee_id and date_from_file == current_date:
                # Έλεγχος καταχώρησης για είσοδο
                if state == "Είσοδος" and entry_time:
                    messagebox.showwarning("Ειδοποίηση", "Έχετε ήδη καταχωρηθεί για είσοδο σήμερα.")
                    entry.delete(0, tk.END)
                    clear_canvas(canvas)
                    return

                # Έλεγχος καταχώρησης για έξοδο
                if state == "Έξοδος" and exit_time:
                    messagebox.showwarning("Ειδοποίηση", "Έχετε ήδη καταχωρηθεί για έξοδο σήμερα.")
                    exit_entry.delete(0, tk.END)
                    clear_canvas(canvas)
                    return

                # Αποθήκευση της υπογραφής
                signature_url = save_signature_image(employee_id, canvas, state)
                if signature_url is None:
                    return

                # Ενημέρωση εγγραφής για είσοδο ή έξοδο
                if state == "Είσοδος" and not entry_time:
                    sheet.cell(row=row_index, column=4, value=current_time)
                    sheet.cell(row=row_index, column=5, value=signature_url)
                elif state == "Έξοδος" and not exit_time:
                    sheet.cell(row=row_index, column=6, value=current_time)
                    sheet.cell(row=row_index, column=7, value=signature_url)

                # Αποθήκευση αλλαγών
                workbook.save(f'employee_data_{date_str}.xlsx')
                messagebox.showinfo("Επιτυχία", f"Ο υπάλληλος {employee['name']} καταχωρήθηκε με επιτυχία!")
                entry.delete(0, tk.END)
                exit_entry.delete(0, tk.END)
                clear_canvas(canvas)
                return

        # Αν δεν βρέθηκε καταχώρηση, προσθήκη νέας
        signature_url = save_signature_image(employee_id, canvas, state)
        if signature_url is None:
            return

        if state == "Είσοδος":
            sheet.append([employee_id, employee['name'], current_date, current_time, signature_url, "", ""])
        elif state == "Έξοδος":
            sheet.append([employee_id, employee['name'], current_date, "", "", current_time, signature_url])

        # Αποθήκευση αλλαγών
        workbook.save(f'employee_data_{date_str}.xlsx')
        messagebox.showinfo("Επιτυχία", f"Ο υπάλληλος {employee['name']} καταχωρήθηκε με επιτυχία!")
        entry.delete(0, tk.END)
        exit_entry.delete(0, tk.END)
        clear_canvas(canvas)
    else:
        messagebox.showerror("Σφάλμα", "Ο αριθμός μητρώου δεν βρέθηκε. Παρακαλώ δοκιμάστε ξανά.")
        entry.delete(0, tk.END)
        exit_entry.delete(0, tk.END)
        clear_canvas(canvas)


# Λειτουργία για τη φόρτωση της λίστας υπαλλήλων από το Excel αρχείο
def load_employee_list():
    global employees
    employees = []

    # Έλεγχος αν υπάρχει το αρχείο υπαλλήλων
    if os.path.exists('employees_list.xlsx'):
        workbook = load_workbook('employees_list.xlsx')
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Από τη δεύτερη σειρά και κάτω
            employees.append({"id": str(row[0]).strip(), "name": str(row[1]).strip() if row[1] else ""})
    else:
        messagebox.showerror("Σφάλμα", "Το αρχείο 'employees_list.xlsx' δεν βρέθηκε. Παρακαλώ βεβαιωθείτε ότι υπάρχει.")
        window.quit()
        sys.exit()  # Κλείνει εντελώς το πρόγραμμα





# Δημιουργία γραφικού παραθύρου
window = tk.Tk()
window.title("Καταχώρηση Αριθμού Μητρώου Υπαλλήλου")
window.geometry("675x400")

# Δημιουργία Frame χωρίς προκαθορισμένο πλάτος και ύψος
frame = Frame(window, bd=11)
frame.pack(expand=True, fill=tk.BOTH)  # Επέκταση για να γεμίσει το διαθέσιμο χώρο
#Φόρτωσε την λίστα των υπαλλήλων
load_employee_list()

# Συνάρτηση για την ενεργοποίηση/απενεργοποίηση στοιχείων
def toggle_entry_exit(state):
    if state == "Είσοδος":
        # Καθαρισμός των πεδίων της εξόδου
        exit_entry.delete(0, tk.END)
        exit_canvas.delete("all")
        # Καθαρισμός των πεδίων της εισόδου
        entry.delete(0, tk.END)
        entry_canvas.delete("all")

        entry_canvas.config(bg="white")
        exit_canvas.config(bg="SystemButtonFace")

        exit_entry.config(state="disabled")
        entry.config(state="normal")

        submit_entry_button.config(state="normal")
        submit_exit_button.config(state="disabled")

        entry_button.config(state="disabled")
        exit_button.config(state="normal")

        entry_canvas.bind("<B1-Motion>", paint_entry)
        entry_canvas.bind("<ButtonRelease-1>", reset_position)
        exit_canvas.unbind("<B1-Motion>")
        exit_canvas.unbind("<ButtonRelease-1>")

    elif state == "Έξοδος":
        # Καθαρισμός των πεδίων της εξόδου
        exit_entry.delete(0, tk.END)
        exit_canvas.delete("all")
        # Καθαρισμός των πεδίων της εισόδου
        entry.delete(0, tk.END)
        entry_canvas.delete("all")

        entry_canvas.config(bg="SystemButtonFace")
        exit_canvas.config(bg="white")

        exit_entry.config(state="normal")
        entry.config(state="disabled")

        submit_entry_button.config(state="disabled")
        submit_exit_button.config(state="normal")

        entry_button.config(state="normal")
        exit_button.config(state="disabled")

        entry_canvas.unbind("<B1-Motion>")
        entry_canvas.unbind("<ButtonRelease-1>")
        exit_canvas.bind("<B1-Motion>", paint_exit)
        exit_canvas.bind("<ButtonRelease-1>", exit_reset_position)



# Κουμπιά Εισόδου και Εξόδου
entry_button = tk.Button(frame, text="Είσοδος", command=lambda: toggle_entry_exit("Είσοδος"))
entry_button.grid(row=0, column=0, padx=10, pady=10)

exit_button = tk.Button(frame, text="Έξοδος", command=lambda: toggle_entry_exit("Έξοδος"))
exit_button.grid(row=0, column=1, padx=10, pady=10)


# Πεδίο εισαγωγής για την είσοδο
entry_label = tk.Label(frame, text="Αριθμός Μητρώου Εισόδου:")
entry_label.grid(row=1, column=0, padx=10, pady=5)

entry = tk.Entry(frame, width=30)
entry.grid(row=1, column=0, padx=10, pady=5)



# Πεδίο εισαγωγής για την έξοδο
exit_label = tk.Label(frame, text="Αριθμός Μητρώου Εξόδου:")
exit_label.grid(row=1, column=1, padx=10, pady=5)

exit_entry = tk.Entry(frame, width=30, state="disabled")
exit_entry.grid(row=1, column=1, padx=10, pady=5)

# Καμβάδες για την υπογραφή
entry_canvas = tk.Canvas(frame, bg="white", width=300, height=150, state="normal")
entry_canvas.grid(row=2, column=0, padx=10, pady=10)

exit_canvas = tk.Canvas(frame, bg="SystemButtonFace", width=300, height=150, state="disabled")
exit_canvas.grid(row=2, column=1, padx=10, pady=10)
'''
# Χειρισμός του καμβά για την υπογραφή
def paint_entry(event):
    x1, y1 = (event.x - 1), (event.y - 1)
    x2, y2 = (event.x + 1), (event.y + 1)
    entry_canvas.create_oval(x1, y1, x2, y2, fill="black", width=2)

entry_canvas.bind("<B1-Motion>", paint_entry)

# Χειρισμός του καμβά για την υπογραφή
def paint_exit(event):
    x1, y1 = (event.x - 1), (event.y - 1)
    x2, y2 = (event.x + 1), (event.y + 1)
    exit_canvas.create_oval(x1, y1, x2, y2, fill="black", width=2)

#exit_canvas.bind("<B1-Motion>", paint_exit)
'''
# Ορισμός μεταβλητών για αποθήκευση της προηγούμενης θέσης
prev_x, prev_y = None, None

# Χειρισμός του καμβά για την υπογραφή
def paint_entry(event):
    global prev_x, prev_y

    # Αν δεν υπάρχει προηγούμενη θέση, αποθηκεύεται η τρέχουσα
    if prev_x is None or prev_y is None:
        prev_x, prev_y = event.x, event.y

    # Δημιουργία γραμμής από την προηγούμενη θέση στην τρέχουσα
    entry_canvas.create_line(prev_x, prev_y, event.x, event.y, fill="black", width=2)

    # Ανανεώνεται η προηγούμενη θέση
    prev_x, prev_y = event.x, event.y

# Όταν αφήνεται το κουμπί του ποντικιού, μηδενίζουμε τις θέσεις
def reset_position(event):
    global prev_x, prev_y
    prev_x, prev_y = None, None

# Συνδέουμε τα γεγονότα με τον καμβά
entry_canvas.bind("<B1-Motion>", paint_entry)
entry_canvas.bind("<ButtonRelease-1>", reset_position)

# Ορισμός μεταβλητών για αποθήκευση της προηγούμενης θέσης
exit_prev_x, exit_prev_y = None, None

# Χειρισμός του καμβά για την υπογραφή
def paint_exit(event):
    global exit_prev_x, exit_prev_y

    # Αν δεν υπάρχει προηγούμενη θέση, αποθηκεύεται η τρέχουσα
    if exit_prev_x is None or exit_prev_y is None:
        exit_prev_x, exit_prev_y = event.x, event.y

    # Δημιουργία γραμμής από την προηγούμενη θέση στην τρέχουσα
    exit_canvas.create_line(exit_prev_x, exit_prev_y, event.x, event.y, fill="black", width=2)

    # Ανανεώνεται η προηγούμενη θέση
    exit_prev_x, exit_prev_y = event.x, event.y

# Όταν αφήνεται το κουμπί του ποντικιού, μηδενίζουμε τις θέσεις
def exit_reset_position(event):
    global exit_prev_x, exit_prev_y
    exit_prev_x, exit_prev_y = None, None

# Συνδέουμε τα γεγονότα με τον καμβά
#exit_canvas.bind("<B1-Motion>", paint_entry)
#exit_canvas.bind("<ButtonRelease-1>", reset_position)

# Κουμπιά Καταχώρησης για Είσοδο και Έξοδο
submit_entry_button = tk.Button(frame, text="Καταχώρηση Εισόδου", command=lambda: save_employee_data(entry_canvas,"Είσοδος"))
submit_entry_button.grid(row=3, column=0, padx=10, pady=10)

submit_exit_button = tk.Button(frame, text="Καταχώρηση Εξόδου", command=lambda: save_employee_data(exit_canvas,"Έξοδος"), state="disabled")
submit_exit_button.grid(row=3, column=1, padx=10, pady=10)

# Ενεργοποίηση του focus στο πεδίο εισαγωγής μόλις φορτώσει το παράθυρο
def on_window_ready(event=None):
    entry.focus_set()

# Χρήση της λειτουργίας bind για να τοποθετήσουμε το focus μόλις το παράθυρο γίνει ορατό
window.bind("<Map>", on_window_ready)

# Έναρξη του κύριου βρόγχου του γραφικού παραθύρου
window.mainloop()