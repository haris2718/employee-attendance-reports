import os
from openpyxl import load_workbook, Workbook
from tkinter import Tk, Label, Entry, Button, messagebox ,  Scale, HORIZONTAL
from tkinter.filedialog import askopenfilenames
import make_Data


def merge_employee_data(date_str):

    # Απόκρυψη του κύριου παραθύρου του Tkinter
    root = Tk()
    root.withdraw()

    # Επιλογή αρχείων για συγχώνευση
    files_to_merge = askopenfilenames(title="Επιλέξτε τα αρχεία για συγχώνευση",
                                      filetypes=[("Excel files", "*.xlsx")],
                                      initialdir=os.getcwd())  # Αρχικός φάκελος

    # Έλεγχος αν επιλέχθηκαν αρχεία
    if not files_to_merge:
        messagebox.showerror("Σφάλμα", f"Δεν επιλέχθηκαν αρχεία για την ημερομηνία {date_str}.")
        return None

    merged_filename = f"employee_data_{date_str}.xlsx"
    merged_workbook = Workbook()
    merged_sheet = merged_workbook.active
    header_written = False

    # Χρήση set για την αποφυγή διπλών εγγραφών
    unique_rows = set()

    for file in files_to_merge:
        workbook = load_workbook(file)
        sheet = workbook.active
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            # Αν δεν έχει γραφτεί ακόμα η κεφαλίδα, γράφουμε την πρώτη γραμμή
            if i == 0 and not header_written:
                merged_sheet.append(row)
                header_written = True
            # Γράφουμε τα δεδομένα από τη 2η γραμμή και μετά αν δεν είναι διπλά
            elif i > 0:
                row_tuple = tuple(row)  # Μετατροπή της γραμμής σε tuple για το set
                if row_tuple not in unique_rows:
                    unique_rows.add(row_tuple)  # Προσθήκη στο set για αποφυγή διπλών εγγραφών
                    merged_sheet.append(row)

    # Αποθήκευση του συγχωνευμένου αρχείου
    merged_workbook.save(merged_filename)
    return merged_filename


def generate_report(date_str,delay):
    # Κλήση της συνάρτησης για συγχώνευση των αρχείων
    merged_file = merge_employee_data(date_str)

    # Ελεγχος αν δημιουργήθηκε το συγχωνευμένο αρχείο
    if merged_file and os.path.exists(merged_file):  # Ελέγχουμε αν το αρχείο υπάρχει
        make_Data.generate_report_pdf_and_excel_with_date(delay,merged_file)
    else:
        messagebox.showerror("Σφάλμα",
                             "Δεν δημιουργήθηκε το συγχωνευμένο αρχείο. Παρακαλώ ελέγξτε τα αρχεία που επιλέξατε.")


# Γραφικό περιβάλλον με Tkinter
def create_gui():
    window = Tk()
    window.title("Αναφορά Υπαλλήλων")
    window.geometry("400x300")

    # Label με δύο σειρές
    Label(window, text="Εισάγετε την ημερομηνία (ΗΗ-ΜΜ-ΕΕΕΕ)\nγια να δημιουργηθεί το προσωρινό αρχείο:").pack(pady=10)
    date_entry = Entry(window)
    date_entry.pack()

    # Προσθήκη μπάρας για επιλογή καθυστέρησης (0 έως 60 λεπτά) και βελτίωση εμφάνισης
    Label(window, text="Επιλέξτε τον χρόνο που μπορεί να \nκαθυστερήσει ένας εργαζόμενος (Λεπτά):").pack(pady=10)
    delay_scale = Scale(window, from_=0, to=60, orient=HORIZONTAL, length=300, sliderlength=30, tickinterval=10,
                        bg="lightgray", fg="black", troughcolor="white")
    delay_scale.set(15)  # Ορισμός της προκαθορισμένης τιμής σε 15
    delay_scale.pack(pady=10)

    def on_generate_report():
        date_str = date_entry.get()
        delay = delay_scale.get()  # Λαμβάνουμε την τιμή της καθυστέρησης από την μπάρα
        if date_str:
            try:
                # Προσπαθεί να επεξεργαστεί και να παράξει την αναφορά
                generate_report(date_str, delay)
            except Exception as e:
                messagebox.showerror("Σφάλμα", f"Παρουσιάστηκε σφάλμα: {str(e)}")

    Button(window, text="Παραγωγή Αναφοράς", command=on_generate_report).pack(pady=10)

    # Συνάρτηση για το κλείσιμο του παραθύρου
    def on_closing():
        if messagebox.askokcancel("Έξοδος", "Θέλετε σίγουρα να κλείσετε το πρόγραμμα;"):
            window.destroy()  # Κλείνει το παράθυρο και τερματίζει το mainloop
            os._exit(0)  # Τερματίζει πλήρως το πρόγραμμα χωρίς να αφήνει διεργασίες ανοιχτές

    # Ορισμός της συνάρτησης που θα καλείται όταν πατιέται το κουμπί κλεισίματος "Χ"
    window.protocol("WM_DELETE_WINDOW", on_closing)

    window.mainloop()

if __name__ == "__main__":
    create_gui()